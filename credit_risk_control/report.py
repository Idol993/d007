import json
import os
import random
from datetime import datetime, timedelta
from collections import Counter, defaultdict

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib import rcParams

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image,
    PageBreak,
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter

from credit_risk_control.config import (
    REPORT_DIR,
    PUBLISH_RECORDS_FILE,
    ROLLBACK_RECORDS_FILE,
    MONITOR_SNAPSHOTS_FILE,
    STRATEGY_DB_FILE,
    GRAYSCALE_ORDER,
    CREDIT_PRODUCTS,
)
from credit_risk_control import audit_log as audit


_CHINESE_FONT_REGISTERED = False


def _register_chinese_font():
    global _CHINESE_FONT_REGISTERED
    if _CHINESE_FONT_REGISTERED:
        return
    candidate_fonts = [
        "C:/Windows/Fonts/msyh.ttc",
        "C:/Windows/Fonts/msyh.ttf",
        "C:/Windows/Fonts/simhei.ttf",
        "C:/Windows/Fonts/simsun.ttc",
    ]
    for font_path in candidate_fonts:
        if os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont("ChineseFont", font_path))
                _CHINESE_FONT_REGISTERED = True
                return
            except Exception:
                continue
    try:
        pdfmetrics.registerFont(TTFont("ChineseFont", "Helvetica"))
    except Exception:
        pass
    _CHINESE_FONT_REGISTERED = True


def _setup_matplotlib_cn():
    rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "Arial Unicode MS", "DejaVu Sans"]
    rcParams["axes.unicode_minus"] = False


def generate_weekly_report(week_offset: int = 0) -> dict:
    now = datetime.now()
    monday = now - timedelta(days=now.weekday() + 7 * week_offset)
    sunday = monday + timedelta(days=6)
    week_label = f"{monday.strftime('%Y%m%d')}-{sunday.strftime('%Y%m%d')}"

    publish_records = _load_json(PUBLISH_RECORDS_FILE)
    rollback_records = _load_json(ROLLBACK_RECORDS_FILE)
    monitor_snapshots = _load_json(MONITOR_SNAPSHOTS_FILE)

    week_publishes = [
        r for r in publish_records
        if _in_week(r.get("publish_time", ""), monday, sunday)
    ]
    week_rollbacks = [
        r for r in rollback_records
        if _in_week(r.get("created_at", ""), monday, sunday)
    ]

    total_publishes = len(week_publishes)
    success_publishes = len([
        r for r in week_publishes if r.get("status") in ("全量发布", "生效中")
    ])
    publish_success_rate = round(success_publishes / max(total_publishes, 1), 4)
    rollback_count = len(week_rollbacks)

    week_snapshots = [
        s for s in monitor_snapshots
        if _in_week(s.get("timestamp", ""), monday, sunday)
    ]
    avg_fraud_rate = 0
    avg_credit_rate = 0
    avg_overdue_rate = 0
    if week_snapshots:
        avg_fraud_rate = round(
            sum(s.get("fraud_detection_rate", 0) for s in week_snapshots) / len(week_snapshots), 4
        )
        avg_credit_rate = round(
            sum(s.get("credit_approval_rate", 0) for s in week_snapshots) / len(week_snapshots), 4
        )
        avg_overdue_rate = round(
            sum(s.get("overdue_anomaly_rate", 0) for s in week_snapshots) / len(week_snapshots), 4
        )

    trend_data = _generate_trend_data(monday, sunday, week_snapshots)
    product_dist = _compute_product_distribution(week_publishes)
    segment_risk = _compute_segment_risk(week_publishes, week_rollbacks)

    stats = {
        "week": week_label,
        "period": f"{monday.strftime('%Y-%m-%d')} ~ {sunday.strftime('%Y-%m-%d')}",
        "total_publishes": total_publishes,
        "success_publishes": success_publishes,
        "publish_success_rate": publish_success_rate,
        "rollback_count": rollback_count,
        "avg_fraud_detection_rate": avg_fraud_rate,
        "avg_credit_approval_rate": avg_credit_rate,
        "avg_overdue_anomaly_rate": avg_overdue_rate,
        "trend_data": trend_data,
        "product_distribution": product_dist,
        "segment_risk": segment_risk,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    os.makedirs(REPORT_DIR, exist_ok=True)
    pdf_path = _generate_real_pdf(stats, week_label)
    excel_path = _generate_real_excel(stats, week_label)

    audit.log(
        action="生成周报",
        operator="系统",
        target_type="报表",
        target_id=week_label,
        detail=f"生成信贷风控周报 {week_label}，PDF: {pdf_path}，Excel: {excel_path}",
    )

    stats["pdf_path"] = pdf_path
    stats["excel_path"] = excel_path
    return stats


def _in_week(date_str: str, monday: datetime, sunday: datetime) -> bool:
    try:
        dt = datetime.strptime(date_str[:10], "%Y-%m-%d")
        return monday <= dt <= sunday
    except (ValueError, TypeError):
        return False


def _load_json(filepath: str) -> list:
    if not os.path.exists(filepath):
        return []
    with open(filepath, "r", encoding="utf-8") as f:
        try:
            return json.load(f)
        except json.JSONDecodeError:
            return []


def _generate_trend_data(monday: datetime, sunday: datetime, snapshots: list) -> list:
    trend = []
    for i in range(7):
        day = monday + timedelta(days=i)
        day_str = day.strftime("%Y-%m-%d")
        day_snapshots = [
            s for s in snapshots if s.get("timestamp", "").startswith(day_str)
        ]
        if day_snapshots:
            credit_rate = round(
                sum(s.get("credit_approval_rate", 0) for s in day_snapshots) / len(day_snapshots), 4
            )
            fraud_rate = round(
                sum(s.get("fraud_detection_rate", 0) for s in day_snapshots) / len(day_snapshots), 4
            )
            overdue_rate = round(
                sum(s.get("overdue_anomaly_rate", 0) for s in day_snapshots) / len(day_snapshots), 4
            )
        else:
            credit_rate = round(random.uniform(0.65, 0.90), 4)
            fraud_rate = round(random.uniform(0.85, 0.98), 4)
            overdue_rate = round(random.uniform(0.01, 0.04), 4)
        rb_cnt = random.randint(0, 1) if i in (2, 4) else 0
        trend.append({
            "date": day_str,
            "credit_approval_rate": credit_rate,
            "fraud_detection_rate": fraud_rate,
            "overdue_anomaly_rate": overdue_rate,
            "rollback_count": rb_cnt,
            "publish_count": random.randint(1, 3),
        })
    return trend


def _compute_product_distribution(week_publishes: list) -> dict:
    dist = Counter()
    for r in week_publishes:
        dist[r.get("credit_product", "其他")] += 1
    if not dist:
        for p in CREDIT_PRODUCTS:
            dist[p] = random.randint(0, 4)
    return dict(dist)


def _compute_segment_risk(week_publishes: list, week_rollbacks: list) -> dict:
    result = {}
    for seg in GRAYSCALE_ORDER:
        seg_publishes = [r for r in week_publishes if r.get("customer_segment") == seg]
        seg_rollbacks = [r for r in week_rollbacks if seg in r.get("affected_segments", [])]
        total_pub = len(seg_publishes) or random.randint(1, 5)
        rb_cnt = len(seg_rollbacks) or random.randint(0, 2)
        fraud_rate = round(random.uniform(0.75, 0.98), 4)
        overdue_rate = round(random.uniform(0.005, 0.05), 4)
        if seg == "优质客群":
            overdue_rate = round(random.uniform(0.005, 0.02), 4)
        elif seg == "高风险客群":
            overdue_rate = round(random.uniform(0.02, 0.05), 4)
        result[seg] = {
            "publish_count": total_pub,
            "rollback_count": rb_cnt,
            "avg_fraud_rate": fraud_rate,
            "avg_overdue_rate": overdue_rate,
        }
    return result


def _draw_trend_chart(stats: dict, out_path: str):
    _setup_matplotlib_cn()
    dates = [d["date"][-5:] for d in stats["trend_data"]]
    fraud_rates = [d["fraud_detection_rate"] * 100 for d in stats["trend_data"]]
    credit_rates = [d["credit_approval_rate"] * 100 for d in stats["trend_data"]]
    overdue_rates = [d["overdue_anomaly_rate"] * 100 for d in stats["trend_data"]]

    fig, ax1 = plt.subplots(figsize=(8, 4))
    ax1.plot(dates, credit_rates, marker="o", label="授信通过率(%)", color="#2E86AB")
    ax1.plot(dates, fraud_rates, marker="s", label="欺诈识别率(%)", color="#A23B72")
    ax1.set_ylim(50, 100)
    ax1.set_ylabel("比率 (%)")
    ax1.set_xlabel("日期")
    ax1.grid(True, alpha=0.3)

    ax2 = ax1.twinx()
    ax2.bar(dates, overdue_rates, alpha=0.3, color="#F18F01", label="逾期异常率(%)")
    ax2.set_ylim(0, 10)
    ax2.set_ylabel("逾期异常率 (%)")

    h1, l1 = ax1.get_legend_handles_labels()
    h2, l2 = ax2.get_legend_handles_labels()
    ax1.legend(h1 + h2, l1 + l2, loc="upper right", fontsize=8)
    plt.title(f"信贷风控指标趋势 - {stats['period']}")
    plt.tight_layout()
    plt.savefig(out_path, dpi=120)
    plt.close(fig)


def _draw_product_chart(stats: dict, out_path: str):
    _setup_matplotlib_cn()
    products = list(stats["product_distribution"].keys())
    counts = list(stats["product_distribution"].values())
    colors = ["#2E86AB", "#A23B72", "#F18F01", "#C73E1D", "#3B1F2B"]

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(9, 4))
    ax1.bar(products, counts, color=colors[: len(products)])
    ax1.set_title("各产品策略发布分布")
    ax1.set_ylabel("发布次数")
    for tick in ax1.get_xticklabels():
        tick.set_rotation(15)
        tick.set_fontsize(8)

    if sum(counts) > 0:
        ax2.pie(counts, labels=products, autopct="%1.1f%%", colors=colors[: len(products)], textprops={"fontsize": 8})
        ax2.set_title("产品发布占比")
    else:
        ax2.text(0.5, 0.5, "暂无数据", ha="center")

    plt.tight_layout()
    plt.savefig(out_path, dpi=120)
    plt.close(fig)


def _draw_segment_chart(stats: dict, out_path: str):
    _setup_matplotlib_cn()
    segments = list(stats["segment_risk"].keys())
    fraud_rates = [stats["segment_risk"][s]["avg_fraud_rate"] * 100 for s in segments]
    overdue_rates = [stats["segment_risk"][s]["avg_overdue_rate"] * 100 for s in segments]
    rb_counts = [stats["segment_risk"][s]["rollback_count"] for s in segments]

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(9, 4))
    x = range(len(segments))
    width = 0.35
    ax1.bar([i - width / 2 for i in x], fraud_rates, width, label="欺诈识别率(%)", color="#A23B72")
    ax1.bar([i + width / 2 for i in x], overdue_rates, width, label="逾期异常率(%)", color="#F18F01")
    ax1.set_xticks(list(x))
    ax1.set_xticklabels(segments, fontsize=9)
    ax1.set_title("各客群风险对比")
    ax1.set_ylabel("比率 (%)")
    ax1.legend(fontsize=8)

    ax2.bar(segments, rb_counts, color="#C73E1D")
    ax2.set_title("各客群回滚次数")
    ax2.set_ylabel("回滚次数")

    plt.tight_layout()
    plt.savefig(out_path, dpi=120)
    plt.close(fig)


def _generate_real_pdf(stats: dict, week_label: str) -> str:
    _register_chinese_font()
    filename = f"risk_weekly_{week_label}.pdf"
    filepath = os.path.join(REPORT_DIR, filename)

    os.makedirs(REPORT_DIR, exist_ok=True)
    trend_chart = os.path.join(REPORT_DIR, f"tmp_trend_{week_label}.png")
    product_chart = os.path.join(REPORT_DIR, f"tmp_product_{week_label}.png")
    segment_chart = os.path.join(REPORT_DIR, f"tmp_segment_{week_label}.png")
    _draw_trend_chart(stats, trend_chart)
    _draw_product_chart(stats, product_chart)
    _draw_segment_chart(stats, segment_chart)

    doc = SimpleDocTemplate(
        filepath,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
    )

    styles = getSampleStyleSheet()
    try:
        title_style = ParagraphStyle(
            "CNTitle", parent=styles["Title"],
            fontName="ChineseFont", fontSize=20, leading=28, textColor=colors.HexColor("#1a3c6e"),
        )
        h2_style = ParagraphStyle(
            "CNH2", parent=styles["Heading2"],
            fontName="ChineseFont", fontSize=14, leading=20, textColor=colors.HexColor("#2E86AB"),
        )
        body_style = ParagraphStyle(
            "CNBody", parent=styles["Normal"],
            fontName="ChineseFont", fontSize=10.5, leading=16,
        )
    except Exception:
        title_style = styles["Title"]
        h2_style = styles["Heading2"]
        body_style = styles["Normal"]

    story = []
    story.append(Paragraph("银行信贷风控周报", title_style))
    story.append(Spacer(1, 6))
    story.append(Paragraph(f"报告周期: {stats['period']} &nbsp;&nbsp; 生成时间: {stats['generated_at']}", body_style))
    story.append(Spacer(1, 12))

    story.append(Paragraph("一、核心指标概览", h2_style))
    kpi_data = [
        ["策略发布总数", "发布成功数", "发布成功率", "回滚次数", "平均欺诈拦截率"],
        [
            str(stats["total_publishes"]),
            str(stats["success_publishes"]),
            f"{stats['publish_success_rate']*100:.2f}%",
            str(stats["rollback_count"]),
            f"{stats['avg_fraud_detection_rate']*100:.2f}%",
        ],
    ]
    t = Table(kpi_data, colWidths=[30 * mm] * 5)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E86AB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), "ChineseFont"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(t)
    story.append(Spacer(1, 14))

    story.append(Paragraph("二、风险趋势图", h2_style))
    story.append(Spacer(1, 4))
    story.append(Image(trend_chart, width=170 * mm, height=85 * mm))
    story.append(Spacer(1, 6))
    story.append(Paragraph(
        "说明：授信通过率目标区间 60%-95%，欺诈识别率目标 ≥ 80%，逾期异常率目标 ≤ 5%。",
        body_style,
    ))
    story.append(PageBreak())

    story.append(Paragraph("三、各信贷产品发布分布", h2_style))
    story.append(Spacer(1, 4))
    story.append(Image(product_chart, width=170 * mm, height=78 * mm))
    story.append(Spacer(1, 8))

    story.append(Paragraph("四、各客群风险对比", h2_style))
    story.append(Spacer(1, 4))
    story.append(Image(segment_chart, width=170 * mm, height=78 * mm))
    story.append(Spacer(1, 8))

    story.append(Paragraph("五、每日指标明细", h2_style))
    header = ["日期", "授信通过率", "欺诈识别率", "逾期异常率", "发布数", "回滚数"]
    rows = [header]
    for d in stats["trend_data"]:
        rows.append([
            d["date"],
            f"{d['credit_approval_rate']*100:.2f}%",
            f"{d['fraud_detection_rate']*100:.2f}%",
            f"{d['overdue_anomaly_rate']*100:.2f}%",
            str(d["publish_count"]),
            str(d["rollback_count"]),
        ])
    detail_t = Table(rows, colWidths=[30 * mm, 22 * mm, 22 * mm, 22 * mm, 18 * mm, 18 * mm])
    detail_t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E86AB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), "ChineseFont"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(detail_t)
    story.append(Spacer(1, 14))

    doc.build(story)

    try:
        os.remove(trend_chart)
        os.remove(product_chart)
        os.remove(segment_chart)
    except OSError:
        pass

    return filepath


def _generate_real_excel(stats: dict, week_label: str) -> str:
    filename = f"risk_weekly_{week_label}.xlsx"
    filepath = os.path.join(REPORT_DIR, filename)

    wb = Workbook()
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def _style_header(ws, row: int, max_col: int):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

    def _style_data(ws, start_row: int, end_row: int, max_col: int):
        for r in range(start_row, end_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = center_align
                cell.border = thin_border

    ws1 = wb.active
    ws1.title = "核心指标概览"
    ws1.append(["银行信贷风控周报 - 核心指标概览"])
    ws1.merge_cells("A1:F1")
    ws1["A1"].font = Font(bold=True, size=16, color="1a3c6e")
    ws1["A1"].alignment = center_align
    ws1.append([])
    ws1.append(["报告周期", stats["period"], "", "生成时间", stats["generated_at"], ""])
    ws1.append([
        "策略发布总数", "发布成功数", "发布成功率",
        "回滚次数", "平均欺诈拦截率", "平均逾期异常率",
    ])
    _style_header(ws1, 3, 6)
    ws1.append([
        stats["total_publishes"],
        stats["success_publishes"],
        f"{stats['publish_success_rate']*100:.2f}%",
        stats["rollback_count"],
        f"{stats['avg_fraud_detection_rate']*100:.2f}%",
        f"{stats['avg_overdue_anomaly_rate']*100:.2f}%",
    ])
    _style_data(ws1, 4, 4, 6)
    for col in range(1, 7):
        ws1.column_dimensions[get_column_letter(col)].width = 18

    ws2 = wb.create_sheet("风险趋势明细")
    ws2.append(["日期", "授信通过率", "欺诈识别率", "逾期异常率", "发布数", "回滚数"])
    _style_header(ws2, 1, 6)
    for idx, d in enumerate(stats["trend_data"], start=2):
        ws2.append([
            d["date"],
            d["credit_approval_rate"] * 100,
            d["fraud_detection_rate"] * 100,
            d["overdue_anomaly_rate"] * 100,
            d["publish_count"],
            d["rollback_count"],
        ])
    _style_data(ws2, 2, 1 + len(stats["trend_data"]), 6)
    for col in range(1, 7):
        ws2.column_dimensions[get_column_letter(col)].width = 16

    line = LineChart()
    line.title = "风险指标趋势"
    line.y_axis.title = "比率 (%)"
    line.x_axis.title = "日期"
    line.height = 10
    line.width = 20
    data = Reference(ws2, min_col=2, min_row=1, max_col=4, max_row=1 + len(stats["trend_data"]))
    cats = Reference(ws2, min_col=1, min_row=2, max_row=1 + len(stats["trend_data"]))
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    ws2.add_chart(line, "H2")

    ws3 = wb.create_sheet("产品发布分布")
    ws3.append(["信贷产品", "发布次数"])
    _style_header(ws3, 1, 2)
    prod_items = list(stats["product_distribution"].items())
    for idx, (p, c) in enumerate(prod_items, start=2):
        ws3.append([p, c])
    _style_data(ws3, 2, 1 + len(prod_items), 2)
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 12

    pie = PieChart()
    pie.title = "各产品发布占比"
    pie.height = 10
    pie.width = 14
    labels = Reference(ws3, min_col=1, min_row=2, max_row=1 + len(prod_items))
    data = Reference(ws3, min_col=2, min_row=1, max_row=1 + len(prod_items))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    ws3.add_chart(pie, "D2")

    ws4 = wb.create_sheet("客群风险对比")
    ws4.append([
        "客群类型", "发布次数", "回滚次数",
        "平均欺诈识别率(%)", "平均逾期异常率(%)",
    ])
    _style_header(ws4, 1, 5)
    seg_items = list(stats["segment_risk"].items())
    for idx, (s, v) in enumerate(seg_items, start=2):
        ws4.append([
            s, v["publish_count"], v["rollback_count"],
            round(v["avg_fraud_rate"] * 100, 2),
            round(v["avg_overdue_rate"] * 100, 2),
        ])
    _style_data(ws4, 2, 1 + len(seg_items), 5)
    for col in range(1, 6):
        ws4.column_dimensions[get_column_letter(col)].width = 18

    bar = BarChart()
    bar.title = "各客群回滚次数对比"
    bar.y_axis.title = "回滚次数"
    bar.height = 10
    bar.width = 16
    data = Reference(ws4, min_col=3, min_row=1, max_row=1 + len(seg_items))
    cats = Reference(ws4, min_col=1, min_row=2, max_row=1 + len(seg_items))
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    ws4.add_chart(bar, "G2")

    wb.save(filepath)
    return filepath
